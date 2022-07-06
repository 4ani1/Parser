
import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl import load_workbook
from collections import OrderedDict
from openpyxl.styles import PatternFill
import datetime

sub_data = datetime.date.today()
sub_time = datetime.datetime.today()
date = f'{sub_data.day}.0{sub_data.month} {sub_time.hour}.{sub_time.minute}'
print(date)

file_name = f'Отчет {date}.xlsx'  # Наименование таблицы

const_class = []
items = []
current_color = ''

url = 'https://mck-ktits.ru/rank/'
response = requests.get(url)
soup = BeautifulSoup(response.text, 'lxml')

yerr = []
const_for_r = []
scaning_tittle = soup.find_all('table')
for index in range(3, len(scaning_tittle)):
    yerr.append(re.findall(r'(?<=tr class=").*?(?=">)', str(scaning_tittle[index])))
listNumbers = list(set(yerr[0]))
for index in range(0, len(listNumbers)):
    if listNumbers[index] == 'R3' or listNumbers[index] == 'R4':
        const_class.append('')
    else:
        const_class.append(listNumbers[index])
for value in const_class:
    if value != '':
        const_for_r.append(value)
items = soup.find_all('tr', class_=[str(const_for_r[0]), str(const_for_r[1])])
print(str(const_for_r[0]), str(const_for_r[1]))

sub_spec_name = ''
sub_type_fic = ''
sub_count_places = ''

for value in const_for_r:
    if value != 'R1':
        sub_spec_name = str(value+'C0')
        sub_type_fic = str(value+'C2')
        sub_count_places = str(value+'C3')

name_speciality = ['Инфокоммуникационные сети и системы связи','Информационные системы и программирование (Программист)',
                   'Информационные системы и программирование (Разработчик веб и мультимедийных приложений)',
                   'Компьютерные системы и комплексы', 'Обеспечение информационной безопасности автоматизированных систем',
                   'Сетевое и системное администрирование', 'Почтовая связь']

reduct_speciality = ['ИССС','ИСП П','ИСП ВЕБ','КСК', 'ОИБАС', 'ССА', 'ПЧ']

type_financing = ['Коммерческое финансирование','Бюджетное финансирование']

reduct_financing = ['КФ','БФ']

colum_name = ['№','ФИО','Док. об образовании','Средний балл аттестата']

spec_name = None
type_fic = None
count_places = None

fio = None
doc = None
grade = None

current_list = 'Sheet'

count = 1
count_list = 2
all_student = []

sub_grade = 0

wb = openpyxl.Workbook()
wb.save(file_name)


def short_spec_name(spec_name, type_fic):  # Функция для сокрощения названий
    for index in range(0, len(name_speciality)):  # Сокращаем название специальности
        if spec_name == name_speciality[index]:
            spec_name = reduct_speciality[index]
    for index in range(0, len(type_financing)):  # Сокращаем тип обучения
        if type_fic == type_financing[index]:
            type_fic = reduct_financing[index]
    return spec_name, type_fic


def setter_data_table(items):
    global count_places, type_fic, spec_name, number, fio, grade, count, current_list, count_list, current_color, sub_grade

    colum_name = ['№', 'ФИО', 'Док. об образовании', 'Средний балл аттестата']

    wb = load_workbook(file_name)
    sheet = wb[current_list]

    places_current_spec = 0
    student_count = 1

    sub_all_grade = []

    flag = True

    for n, i in enumerate(items, start=1):
        if i.find('td', class_='R5C0') is not None:  # Название
            spec_name = re.findall(r'(?<=">).*?(?=</td>)', str(i.find('td', class_='R5C0')))[0]

        if i.find('td', class_=sub_spec_name) is not None:  # Название для еблана
            spec_name = re.findall(r'(?<=">).*?(?=</td>)', str(i.find('td', class_=sub_spec_name)))[0]

        if i.find('td', class_='R5C2') is not None:  # Тип финансирования
            type_fic = re.findall(r'(?<=">).*?(?=</td>)', str(i.find('td', class_='R5C2')))[0]
        if i.find('td', class_=sub_type_fic) is not None:  # Тип финансирования из-за ебланаской верстки
            type_fic = re.findall(r'(?<=">).*?(?=</td>)', str(i.find('td', class_=sub_type_fic)))[0]

        if i.find('td', class_='R5C3') is not None:  # Места
            count_places = re.findall(r'(?<=">).*?(?=</td>)', str(i.find('td', class_='R5C3')))[0]
            count = 1
            count_list = 2
        else:
            if i.find('td', class_=sub_count_places) is not None:  # Места
                count_places = re.findall(r'(?<=">).*?(?=</td>)', str(i.find('td', class_=sub_count_places)))[0]
                count = 1
                count_list = 2
            else:
                count_places = None
                type_fic = None
                spec_name = None

        if spec_name is not None and type_fic is not None and count_places is not None: # Это работает когда есть данные
            short_name = short_spec_name(spec_name, type_fic)
            print(f'{short_name[0]} {short_name[1]} {count_places}')
            places_current_spec = int(count_places)
            student_count = 1
            sheet = wb.create_sheet(f'{short_name[0]} {short_name[1]} {count_places}') #  Создаем лист с названием специальности
            if flag:
                del wb['Sheet']
                flag = False
            else:
                for value in sub_all_grade:
                    value = str(value).replace(",", ".")
                    sub_grade = sub_grade + float(value)

                sheet.cell(row=1, column=5).value = 'Общий сред. балл'
                sheet.cell(row=2, column=5).value = str(str(sub_grade / len(sub_all_grade)))[:4]
                sub_grade = 0

            sheet.column_dimensions['A'].width = 5 # Размер столбца с нумерацией
            sheet.column_dimensions['B'].width = 40 # Размер столбца с ФИО
            sheet.column_dimensions['C'].width = 25 # Размер столбца с указанием подонного документа
            sheet.column_dimensions['D'].width = 25 # Размер столбца со средним балом
            sheet.column_dimensions['E'].width = 30 # Размер столбца со общийм средним балом

        if count == 1: # Добавляем верхнюю линию с описанием столбцов
            for index, value in enumerate(colum_name, 1):
                sheet.cell(row=1, column=index).value = value


        number = None
        fio = None
        grade = None
        doc = None

        sub_doc = re.findall(r'(?<="2">).*?(?=</td>)', str(i)[str(i).find('colspan="2">'):str(i).find('colspan="2">')+30]) # Тип документа
        """
         Верхняя строка релозована для поиска элемента отвечающий за документы 
         параметры данной строки полностью совпадают с элементом отвечающий за хранение ФИО
         по этой причине пришлось написать такой костыль
        """
        if len(sub_doc) > 0:
            try:
                doc = sub_doc[0]
            except:
                doc = ''
            sheet.cell(row=count_list, column=3).value = doc
            if doc == 'Копия': # Тут будет серый так как копия и не факт что студент сдаст оригинал
                current_color = '9b9b9b'
            else: # Тут уже оригинал
                if student_count < places_current_spec + 1:
                    student_count = student_count + 1
                    current_color = '98FB98' # Зеленый студент проходит по квоте с оригиналом
                else:
                    current_color = 'ff6161' # Красный студент не прошел квоту с оригиналом

            sheet.cell(row=count_list, column=3).fill = PatternFill(start_color=current_color, end_color=current_color,
                                                                    fill_type='solid') # Красим ячейку документов

        sheet.cell(row=count_list, column=1).value = count

        if i.find('td', class_='R6C1') is not None:  # ФИО
            try:
                fio = re.findall(r'(?<=">).*?(?=</td>)', str(i.find('td', class_='R6C1')))[0]
            except:
                fio = ''
            sheet.cell(row=count_list, column=2).value = fio
            sheet.cell(row=count_list, column=2).fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid') # Красим ячейку ФИО
            sheet.cell(row=count_list, column=1).fill = PatternFill(start_color=current_color, end_color=current_color,
                                                                    fill_type='solid') # Красим ячейку подсчета

        if i.find('td', class_='R6C4') is not None:  # Сред. оценка
            try:
                grade = re.findall(r'(?<=0px;">).*?(?=</span>)', str(i.find('td', class_='R6C4')))[0]
            except IndexError:
                grade = 0
            sub_all_grade.append(grade)
            sheet.cell(row=count_list, column=4).value = grade
            sheet.cell(row=count_list, column=4).fill = PatternFill(start_color=current_color, end_color=current_color,
                                                                    fill_type='solid') # Красим ячейку средней оценки

        if fio is not None and grade is not None:
            print(f'№{count} {fio}, {doc}, {grade}')
            all_student.append(f'{fio}, {doc}, {grade}') # Добовляем всех студентов в один лист
            count_list = count_list + 1 # Добавляем всех студентов в один лист
            count = count + 1 # Увеличиваем счетчик подсчета студентов для специальности

        wb.save(file_name) # Итоги каждой итерации сохраняем


setter_data_table(items)


clear_all_student = list(OrderedDict.fromkeys(all_student))

five_student = []

count_orig = 0
count_cop = 0

for n, i in enumerate(clear_all_student, start=1):
    print(f'№{n} {i}')

for student in clear_all_student:
    if re.search(r'\bОригинал\b', student):
        count_orig = count_orig + 1
    else:
        count_cop = count_cop + 1
    if student[-4:] == '5,00':
        five_student.append(student)

print(f'Кол. ориг: {count_orig}')
print(f'Кол. копий: {count_cop}')
print(len(clear_all_student))

wb = load_workbook(file_name)
sheet = wb.create_sheet('Общая информация')

sheet.column_dimensions['A'].width = 5  # Размер столбца с нумерацией
sheet.column_dimensions['B'].width = 50  # Размер столбца с ФИО
sheet.column_dimensions['C'].width = 10  # Размер столбца с ФИО

sheet.cell(row=2, column=2).value = 'Общее количество абитуриентов:'
sheet.cell(row=3, column=2).value = 'Общее количество оригиналов документов:'
sheet.cell(row=4, column=2).value = 'Общее количество копий документов:'

sheet.cell(row=2, column=3).value = str(len(clear_all_student))
sheet.cell(row=3, column=3).value = str(count_orig)
sheet.cell(row=4, column=3).value = str(count_cop)

sheet.cell(row=6, column=2).value = 'Абитуриенты со сред. баллом 5.00'
sheet.cell(row=7, column=1).value = '№'
sheet.cell(row=7, column=2).value = 'ФИО, Док. об образовании, Средний балл аттестата'

for index in range(1, len(five_student)):
    sheet.cell(row=7+index, column=1).value = index
    sheet.cell(row=7+index, column=2).value = five_student[index]

wb.save(file_name)